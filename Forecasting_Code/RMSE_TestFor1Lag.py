import pandas as pd
import LoadMyModel
import numpy as np
from openpyxl import Workbook, load_workbook
from pathlib import Path

def rmse(predicted_values, real_values):
    predicted_values = np.array(predicted_values)
    real_values = np.array(real_values)
    squared_diff = (predicted_values - real_values) ** 2
    mean_squared_diff = np.mean(squared_diff)
    rmse = np.sqrt(mean_squared_diff)
    return rmse

def mae(predicted_values, real_values):
    predicted_values = np.array(predicted_values)
    real_values = np.array(real_values)
    abs_diff = np.abs(predicted_values - real_values)
    mae = np.mean(abs_diff)
    return mae

def mape(real_values, predicted_values):
    actual = np.array(real_values)
    predicted = np.array(predicted_values)

    # Calculate absolute percentage error for each element
    absolute_percentage_error = np.abs((actual - predicted) / actual)

    # Replace infinite values with NaN and then drop them
    absolute_percentage_error = np.where(np.isfinite(absolute_percentage_error), absolute_percentage_error, np.nan)
    absolute_percentage_error = absolute_percentage_error[~np.isnan(absolute_percentage_error)]

    # Calculate MAPE
    mape = np.mean(absolute_percentage_error) * 100

    return mape



pd.set_option('display.max_columns', None)
PATH = r"https://raw.githubusercontent.com/LGsus20/Weather-Forecasting-Scraping/main/DATASETS/DATASET_Modified_Monthly_2021-2024.csv"
new_data = pd.read_csv(PATH)
MY_VALUES_PATCHTST = []
MY_VALUES_LSTM = []
MY_REAL_VALUES = []

memoryInNLForecast = 216
# starting_row = 26280   # THIS IS THE FIRST HOUR OF Jan/01/2024 IN THE DATASET 2021_2024
starting_row = 26064   # THIS IS THE FIRST HOUR OF Jan/01/2024 IN THE DATASET 2021_2024 taking into account
                         # 216 lags behind

for i in range(len(new_data)-(starting_row + memoryInNLForecast)):
    TEST_VARIABLE = len(new_data)-(starting_row + memoryInNLForecast)
    i += 1
    ending_row = starting_row + memoryInNLForecast + i
    # DATA THE MODEL WILL GET TO PREDICT
    dataForPrediction = new_data.assign(unique_id=np.ones(len(new_data)))
    dataForPrediction = dataForPrediction.iloc[(starting_row+i):ending_row]
    dataForPrediction['ds'] = pd.to_datetime(dataForPrediction['ds'])
    # REAL DATA
    real_data = new_data.assign(unique_id=np.ones(len(new_data)))
    real_data = real_data.iloc[ending_row:(ending_row+1)]
    real_data['ds'] = pd.to_datetime(real_data['ds'])

    # CALL MODEL FOR PREDICTION
    Y_hat_df = LoadMyModel.LoadMyNeuralModel(dataForPrediction)

    # STORE VALUES
    try:
        MY_REAL_VALUES.append(float(real_data["y"].iloc[0]))
        MY_VALUES_PATCHTST.append((float(Y_hat_df["PatchTST"].iloc[0])))
        MY_VALUES_LSTM.append((float(Y_hat_df["LSTM"].iloc[0])))
    except(IndexError):
        break

    print(f"ITERATION: {i}")


excel_file = f'Benchmark_with_starting_row_{starting_row}.xlsx'
file_path = Path(excel_file)

if not file_path.is_file():
    wb = Workbook()
    wb.save(excel_file)
wb = load_workbook(excel_file)
ws = wb.active

ws.append([''])
ws.append(['RMSE LSTM:', rmse(MY_VALUES_LSTM, MY_REAL_VALUES)])
ws.append(['RMSE PatchTST:', rmse(MY_VALUES_PATCHTST, MY_REAL_VALUES)])
ws.append(['MAE LSTM:', mae(MY_VALUES_LSTM, MY_REAL_VALUES)])
ws.append(['MAE PatchTST:', mae(MY_VALUES_PATCHTST, MY_REAL_VALUES)])
ws.append(['MAPE LSTM:', mape(MY_VALUES_LSTM, MY_REAL_VALUES)])
ws.append(['MAPE PatchTST:', mape(MY_VALUES_PATCHTST, MY_REAL_VALUES)])
wb.save(excel_file)